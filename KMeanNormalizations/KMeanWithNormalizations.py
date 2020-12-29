"""
This program is responding to Phase 2: Normalization and Initialization
Min_Max() is responding to Main-Max normalization
z_score() is responding to z-score normalization
K_Mean_Maximin() is responding to K-Mean with Maximin method
K_Mean_rdPartition() is responding to K-Mean with random partition implementation
K_Mean_rdCenters() is responding to  K-mean with randomly selecting centers (phase1)
K_mean() is responding to the general steps after center selecting, Initial SEE, Final SSE, Numver of iterations are updating here
"""
import pandas as pd
import random as rd
import sys
import math
import numpy as np

#rd.seed(0)
# #Load txt file into dataframe
# def load_txt(filename):
#     data = pd.read_csv(filename, sep=" ", skiprows=1, header=None)
#     #data = Min_Max(data)
#     #data = z_score(data)
#     data = data.values
#     return data

#Min-Max normalization
def Min_Max(filename):
    data = pd.read_csv(filename, sep=" ", skiprows=1, header=None)
    #v'=((v-min)/(max-min))*(new_max-new_min)+new_min
    #new_max = 1; new_min=0 here
    #so the equation will be v'=(v-min)/(max-min)
    #data.min() returns the minimum of each column
    #data.max() returns the maximum of each column
    data = (data-data.min())/(data.max()-data.min())
    data = data.fillna(0)
    data = data.values
    return data

#z-score normalization
def z_score(filename):
    # Load txt file into dataframe
    data = pd.read_csv(filename, sep=" ", skiprows=1, header=None)
    #v' = (v-mean)/standard_deviation
    #data.mean() returns the mean of each column
    #data.std() returns the standard deviation of each column
    data = (data-data.mean())/(data.std())
    data = data.fillna(0)
    data = data.values
    return data

#K-Mean with Maximin method
def K_Mean_Maximin(data, clusters, iterations, threshold):
    #create an empty array with NaN values
    centers  = np.empty((clusters,data.shape[1]))
    centers[:] = np.NaN
    #randomly select the first center
    center_first = rd.randrange(0,data.shape[0],1)
    centers[0] = data[center_first,:]
    #loop the all array that contains the centers
    for i in range(clusters):
        # we will fill up the center where is empty
        if(np.isnan(centers[i])).any() == True:
            # greatest distance
            dis_max = 0
            # to track the center
            track_center = 0
            #Look for the point that has the greatest edclidean distance to the nearest previously selected (i-1) centers
            for j in range(data.shape[0]):
                # Eduidance distance;
                # this loop will return the smallest distance between the point to each centers
                for k in range(i):
                    if k == 0:
                        dist_center_x_0 = dist(centers[k],data[j])
                    else:
                        dist_center_x = dist(centers[k],data[j])
                        if (dist_center_x < dist_center_x_0):
                            dist_center_x_0 = dist_center_x
                #assign the center with greatest distance among all
                if (dis_max < dist_center_x_0):
                    dis_max = dist_center_x_0
                    track_center = j
            centers[i] = data[track_center]
    # Add a column to keep track of the cluster of each data point; as fitst column of the dataset
    data = np.c_[[99] * data.shape[0], data]
    initial_sse, final_sse, num_intera = K_Mean(centers, data, clusters, iterations, threshold)
    return initial_sse, final_sse, num_intera

#K-Mean with random partition implementation
def K_Mean_rdPartition(data, clusters, iterations, threshold):
    #Add a column to keep track of the cluster of each data point; as fitst column of the dataset
    data = np.c_[[99]*data.shape[0],data]
    #Baed on the introduction you gave, we tandom initiate the cluster to each data point instead of the random centers of each clusters
    for i in range(data.shape[0]):
        data[i,0] = rd.randrange(0,clusters,1)
    #Calcute the centroids of each clusters
    data_cluster_errs = np.zeros((clusters, data.shape[1] + 1))
    for i in range(data.shape[0]):
        cluster = int(data[i, 0])
        data_cluster_errs[cluster, 0] = cluster
        data_cluster_errs[cluster, 1] += 1
        data_cluster_errs[cluster, 2:data.shape[1] + 1] += data[i, 1:data.shape[1]]
    #Get the centers
    centers = np.zeros((clusters,data.shape[1]-1))
    for i in range(clusters):
        centers[i] = data_cluster_errs[i, 2:data.shape[1] + 1] / data_cluster_errs[i, 1]
     # Call the k_mean() to do the rest as the phase 1
    initial_sse, final_sse, num_intera = K_Mean(centers, data, clusters, iterations, threshold)
    return initial_sse, final_sse, num_intera

def K_Mean_rdCenters(data, clusters, iterations, threshold):
    #Generate random centers for each clusters
    rn = rd.sample(range(0, data.shape[0]), clusters)
    centers = data[rn, :]
    #print(centers.shape[0])
    #Create a dataframe for storing distances of all the instances(potints) to different cluster center
    #The size should be number of instances * number of clusters

    # #Add a 'Cluters' column to original data to indincate which cluster this instance belongs to, second column of the array
    data = np.c_[[99]*data.shape[0],data]
    initial_sse, final_sse, num_intera = K_Mean(centers, data, clusters, iterations, threshold)
    return initial_sse, final_sse, num_intera

def K_Mean(centers,data, clusters, iterations, threshold):
    # Create a list to store the SSE of each iterations
    sum_errors = [99] * iterations

    # #A value to store the improvement of SSE in each iterations
    error_improve = math.inf

    # Create the the 'distance' column to riginal data to indincate the distance of the instance to their cluster center, first column of the array
    data = np.c_[[math.inf] * data.shape[0], data]
    #
    # #A value to keep track of the iteration step (Which iteration it is)
    step = 0

    # #Let jump into the while loop
    # #The while loop will exit if the improvment of SSE small than threshold, or running out of the iterations, or there is no improvememnt in SSE
    while ((error_improve - threshold >= 0) and (iterations - step > 0)):
        # Reset the distance to infinite
        data[:, 0] = [math.inf] * data.shape[0]
        # Calculate all the distance between all the instances to all the cluster centers
        for i in range(data.shape[0]):
            for j in range(clusters):
                # Eduidance distance
                dis = dist(data[i, 2:data.shape[1]], centers[j, :])
                # Store the distance to the cluster_dis dataframe
                # Assign the instance to the closest cluster
                if (dis < data[i][0]):
                    # assign j to the cluster
                    data[i][1] = j
                    # assign closest distance
                    data[i][0] = dis

        data_cluster_errs = np.zeros((clusters, data.shape[1] + 1))
        for i in range(data.shape[0]):
            cluster = int(data[i, 1])
            data_cluster_errs[cluster, 0] += data[i, 0]
            data_cluster_errs[cluster, 1] = cluster
            data_cluster_errs[cluster, 2] += 1
            data_cluster_errs[cluster, 3:data.shape[1] + 1] += data[i, 2:data.shape[1]]

        # Handling empty cluster
        # Check if there is any 0 in the third column of data_cluster errs
        if (np.isin(0, data_cluster_errs[:, 2]).any()) == True:
            for i in range(clusters):
                # Pass if the cluster is not empty(the count is not 0)
                if (np.isin(0, data_cluster_errs[i, 2]).all()) == False:
                    pass
                else:
                    # Sort the data based on the distance with ascend order, then get the last index values, which contributes the most to the overall SSE
                    data = data[np.argsort(data[:, 0])]
                    # Change the cluster name to be the empty cluster
                    data[-1, 1] = i
                    # Change the distance to center to 0
                    data[-1, 0] = 0
            data_cluster_errs = np.zeros((clusters, data.shape[1] + 1))
            for i in range(data.shape[0]):
                cluster = int(data[i, 1])
                data_cluster_errs[cluster, 0] += data[i, 0]
                data_cluster_errs[cluster, 1] = cluster
                data_cluster_errs[cluster, 2] += 1
                data_cluster_errs[cluster, 3:data.shape[1] + 1] += data[i, 2:data.shape[1]]
        # Get the sum of Sum-of-Squared Error
        sum_errors[step] = data_cluster_errs[:, 0].sum()

        # if step == 0:
        #     # Get the initial SSE
        #     print("Initial SSE = {:.4f}".format(sum_errors[step]))
        #     # Skip this part if it is first step since sse[0] = infnite
        if step >= 1:
            # Update the improvement in SSE
            error_improve = (sum_errors[step - 1] - sum_errors[step]) / sum_errors[step - 1]
        #    #Print result
        print("Iteration {}: SSE = {:.4f}".format(step + 1, sum_errors[step]))

        # file.write("Iteration {}: SSE = {:.4f}".format(step+1,sum_errors[step])+'\n')

        # Assign the new centers for next iterations
        for i in range(clusters):
            centers[i] = data_cluster_errs[i, 3:data.shape[1] + 1] / data_cluster_errs[i, 2]
        # Exit the while loop if there is no improvement in SSE
        if error_improve == 0:
            # print("Final SSE = {:.4f}".format(sum_errors[step]))
            # print('Number of Iterations: {}'.format(step))
            return sum_errors[0], sum_errors[step], step
        else:
            step += 1
    # print("Final SSE = {:.4f}".format(sum_errors[step - 1]))
    # print('Number of Iterations: {}'.format(step))
    # Return the lastest SSE
    #return sum_errors[step - 1]
    return sum_errors[0], sum_errors[step - 1], step
# Eduidance distance
def dist(a, b):
    distance = 0
    for i in range(a.shape[0]):
        distance += (a[i]-b[i])*(a[i]-b[i])

    return distance

#Get the best run of k mean with random partition
def Best_run_rdPartition(data, cluster, iteration, thresold,run):
    # best initial SSEs, best final SSEs, and best # iterations
    initial_sse, final_sse, num_intera = K_Mean_rdPartition(data, cluster, iteration, thresold)
    # More runs
    for i in range(run - 1):
        run_initial_sse, run_final_sse, run_num_intera = K_Mean_rdPartition(data, cluster, iteration, thresold)
        # Replace the best_run if other run has better result
        if  initial_sse > run_initial_sse:
            initial_sse = run_initial_sse
        if final_sse > run_final_sse:
            final_sse = run_final_sse
        if num_intera >  run_num_intera:
            num_intera = run_num_intera
    return initial_sse, final_sse, num_intera
#Get the best run of k mean with random centers
def Best_run_rdCenters(data, cluster, iteration, thresold,run):
    # best initial SSEs, best final SSEs, and best # iterations
    initial_sse, final_sse, num_intera = K_Mean_rdCenters(data, cluster, iteration, thresold)
    # More runs
    for i in range(run - 1):
        run_initial_sse, run_final_sse, run_num_intera = K_Mean_rdCenters(data, cluster, iteration, thresold)
        # Replace the best_run if other run has better result
        if  initial_sse > run_initial_sse:
            initial_sse = run_initial_sse
        if final_sse > run_final_sse:
            final_sse = run_final_sse
        if num_intera > run_num_intera:
            num_intera = run_num_intera
    return initial_sse, final_sse, num_intera
#Get the best run of k mean with maximin method
def Best_run_maxmin(data, cluster, iteration, thresold,run):
    # best initial SSEs, best final SSEs, and best # iterations
    initial_sse, final_sse, num_intera = K_Mean_Maximin(data, cluster, iteration, thresold)
    # More runs
    for i in range(run - 1):
        run_initial_sse, run_final_sse, run_num_intera = K_Mean_Maximin(data, cluster, iteration, thresold)
        # Replace the best_run if other run has better result
        if  initial_sse > run_initial_sse:
            initial_sse = run_initial_sse
        if final_sse > run_final_sse:
            final_sse = run_final_sse
        if num_intera >  run_num_intera:
            num_intera = run_num_intera
    return initial_sse, final_sse, num_intera


# def main():
#     #data = load_txt("D:\Python\DataClustering\Phase1\iris_bezdek.txt")
#     data = Min_Max("D:\Python\DataClustering\Phase1\iris_bezdek.txt")
#     num_clusters = 3
#     num_interations = 100
#     threshold = 0.001
#     run = 100
#     initial_sse, final_sse, num_intera = Best_run_maxmin(data, num_clusters, num_interations, threshold, run)
#     print(initial_sse)
#     print(final_sse)
#     print(num_intera)
#main()
#import openpyxl module to open excel file to write data in
import openpyxl
if __name__ == "__main__":
#     #Getting the filename, number of clusters, iterations, threshold, and run from cmd
    print("Script name: %s" % str(sys.argv[0]))
    print("filename: %s" % str(sys.argv[1]))
    print("clusters: %s" % int(sys.argv[2]))
    print("iterations: %s" % int(sys.argv[3]))
    print("thresold: %s" % float(sys.argv[4]))
    print("run: %s" % int(sys.argv[5]))
    print("Saved as: %s" % str(sys.argv[6]))
    saved_sheet = str(sys.argv[6])
    # file = open(saved_file, "w")
    # file.write("Script name: %s" % str(sys.argv[0]) +'\n')
    # file.write("Filename: %s" % str(sys.argv[1]) +'\n')
    # file.write("Clusters: %s" % int(sys.argv[2])+'\n')
    # file.write("Iterations: %s" % int(sys.argv[3])+'\n')
    # file.write("Threshold: %s" % float(sys.argv[4])+'\n')
    # file.write("Number of runs: %s" % int(sys.argv[5])+'\n')
    # file.write("Saved as: %s" % str(sys.argv[6])+'\n')
    wb = openpyxl.load_workbook(filename='D:\Python\DataClustering\Phase2\Phase2.xlsx')
    # Sheets can be added to workbook with the
    # workbook object's create_sheet() method.
    ws = wb[saved_sheet]
    ws['A1'] = 'Filename'
    ws['A2'] = 'Clusters'
    ws['A3'] = 'Max-Iterations'
    ws['A4'] = 'Thresold'
    ws['A5'] = 'Number of Run'

    ws['B1'] = str(sys.argv[1])
    ws['B2'] = int(sys.argv[2])
    ws['B3'] = int(sys.argv[3])
    ws['B4'] = float(sys.argv[4])
    ws['B5'] = int(sys.argv[5])

    ws['A7'] = 'Random Selection'
    ws['A9'] = 'Random Partition'
    ws['A11'] = 'Maximin Method'
    ws['C6'] = 'Best initial SSE'
    ws['D6'] = 'Best final SSE'
    ws['E6'] = 'Best # iterations'
    ws['B7'] = 'Min_Max'
    ws['B8'] = 'Z-score'
    ws['B9'] = 'Min_Max'
    ws['B10'] = 'Z-score'
    ws['B11'] = 'Min_Max'
    ws['B12'] = 'Z-score'
    filename = str(sys.argv[1])
    run = int(sys.argv[5])
    num_clusters = int(sys.argv[2])
    num_interations = int(sys.argv[3])
    threshold = float(sys.argv[4])
    #Min-Max Normalization
    data = Min_Max(filename)
    # Get the best run of k mean with random centers
    initial_sse, final_sse, num_intera = Best_run_rdCenters(data, num_clusters, num_interations, threshold,run)
    ws['C7'] = initial_sse
    ws['D7'] = final_sse
    ws['E7'] = num_intera
    # Get the best run of k mean with random partition
    initial_sse, final_sse, num_intera = Best_run_rdPartition(data, num_clusters, num_interations, threshold,run)
    ws['C9'] = initial_sse
    ws['D9'] = final_sse
    ws['E9'] = num_intera
    # Get the best run of k mean with maximin method
    initial_sse, final_sse, num_intera = Best_run_maxmin(data, num_clusters, num_interations, threshold,run)
    ws['C11'] = initial_sse
    ws['D11'] = final_sse
    ws['E11'] = num_intera
    #z-score Normalization
    data = z_score(filename)
    # Get the best run of k mean with random centers
    initial_sse, final_sse, num_intera = Best_run_rdCenters(data, num_clusters, num_interations, threshold, run)
    ws['C8'] = initial_sse
    ws['D8'] = final_sse
    ws['E8'] = num_intera
    # Get the best run of k mean with random partition
    initial_sse, final_sse, num_intera = Best_run_rdPartition(data, num_clusters, num_interations, threshold, run)
    ws['C10'] = initial_sse
    ws['D10'] = final_sse
    ws['E10'] = num_intera
    # Get the best run of k mean with maximin method
    initial_sse, final_sse, num_intera = Best_run_maxmin(data, num_clusters, num_interations, threshold, run)
    ws['C12'] = initial_sse
    ws['D12'] = final_sse
    ws['E12'] = num_intera

    wb.save('D:\Python\DataClustering\Phase2\Phase2.xlsx')