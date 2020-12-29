"""
This program is responding to Phase 2: Normalization and Initialization
Min_Max() is responding to Main-Max normalization
z_score() is responding to z-score normalization
K_Mean_Maximin() is responding to K-Mean with Maximin method
K_Mean_rdPartition() is responding to K-Mean with random partition implementation
K_Mean_rdCenters() is responding to  K-mean with randomly selecting centers (phase1)
K_mean() is responding to the general steps after center selecting, Initial SEE, Final SSE, Numver of iterations are updating here
"""
import pandas as pd
import random as rd
import sys
import math
import numpy as np
import scipy.misc
#
# rd.seed(5)
# #Load txt file into dataframe
# def load_txt(filename):
#     data = pd.read_csv(filename, sep=" ", skiprows=1, header=None)
#     #data = Min_Max(data)
#     #data = z_score(data)
#     data = data.values
#     return data

#Min-Max normalization
def Min_Max(filename):
    data = pd.read_csv(filename, sep=" ",header=0)
    data_clusters = float(data.columns[2])
    data_clusters = int(np.floor(data_clusters))
    data = pd.read_csv(filename, sep=" ", skiprows=1, header=None)
    data = data.dropna(axis='columns')
    data_labels = data.values[:,data.shape[1]-1]
    #v'=((v-min)/(max-min))*(new_max-new_min)+new_min
    #new_max = 1; new_min=0 here
    #so the equation will be v'=(v-min)/(max-min)
    #data.min() returns the minimum of each column
    #data.max() returns the maximum of each column
    data = data.iloc[:, 0:data.shape[1]-1]
    data = (data-data.min())/(data.max()-data.min())
    data = data.fillna(0)
    data = data.values
    return data_clusters, data, data_labels



def K_Mean_rdCenters(data, clusters, iterations, threshold, data_labels):
    #Generate random centers for each clusters
    rn = rd.sample(range(0, data.shape[0]), clusters)
    centers = data[rn, :]
    #print(centers.shape[0])
    #Create a dataframe for storing distances of all the instances(potints) to different cluster center
    #The size should be number of instances * number of clusters

    # #Add a 'Cluters' column to original data to indincate which cluster this instance belongs to, second column of the array
    data = np.c_[[99]*data.shape[0],data]
    precited_labels = K_Mean(centers, data, clusters, iterations, threshold)
    Jaccard, rand, FM = External_Meature(precited_labels, data_labels)
    return Jaccard, rand, FM

def K_Mean(centers,data, clusters, iterations, threshold):
    # Create a list to store the SSE of each iterations
    sum_errors = [99] * iterations

    # #A value to store the improvement of SSE in each iterations
    error_improve = math.inf

    # Create the the 'distance' column to riginal data to indincate the distance of the instance to their cluster center, first column of the array
    data = np.c_[[math.inf] * data.shape[0], data]
    #
    # #A value to keep track of the iteration step (Which iteration it is)
    step = 0

    # #Let jump into the while loop
    # #The while loop will exit if the improvment of SSE small than threshold, or running out of the iterations, or there is no improvememnt in SSE
    while ((error_improve - threshold >= 0) and (iterations - step > 0)):
        # Reset the distance to infinite
        data[:, 0] = [math.inf] * data.shape[0]
        # Calculate all the distance between all the instances to all the cluster centers
        for i in range(data.shape[0]):
            for j in range(clusters):
                # Eduidance distance
                dis = dist(data[i, 2:data.shape[1]], centers[j, :])
                # Store the distance to the cluster_dis dataframe
                # Assign the instance to the closest cluster
                if (dis < data[i][0]):
                    # assign j to the cluster
                    data[i][1] = j
                    # assign closest distance
                    data[i][0] = dis

        data_cluster_errs = np.zeros((clusters, data.shape[1] + 1))
        for i in range(data.shape[0]):
            cluster = int(data[i, 1])
            data_cluster_errs[cluster, 0] += data[i, 0]
            data_cluster_errs[cluster, 1] = cluster
            data_cluster_errs[cluster, 2] += 1
            data_cluster_errs[cluster, 3:data.shape[1] + 1] += data[i, 2:data.shape[1]]

        # Handling empty cluster
        # Check if there is any 0 in the third column of data_cluster errs
        if (np.isin(0, data_cluster_errs[:, 2]).any()) == True:
            for i in range(clusters):
                # Pass if the cluster is not empty(the count is not 0)
                if (np.isin(0, data_cluster_errs[i, 2]).all()) == False:
                    pass
                else:
                    # Sort the data based on the distance with ascend order, then get the last index values, which contributes the most to the overall SSE
                    data = data[np.argsort(data[:, 0])]
                    # Change the cluster name to be the empty cluster
                    data[-1, 1] = i
                    # Change the distance to center to 0
                    data[-1, 0] = 0
            data_cluster_errs = np.zeros((clusters, data.shape[1] + 1))
            for i in range(data.shape[0]):
                cluster = int(data[i, 1])
                data_cluster_errs[cluster, 0] += data[i, 0]
                data_cluster_errs[cluster, 1] = cluster
                data_cluster_errs[cluster, 2] += 1
                data_cluster_errs[cluster, 3:data.shape[1] + 1] += data[i, 2:data.shape[1]]
        # Get the sum of Sum-of-Squared Error
        sum_errors[step] = data_cluster_errs[:, 0].sum()

        # if step == 0:
        #     # Get the initial SSE
        #     print("Initial SSE = {:.4f}".format(sum_errors[step]))
        #     # Skip this part if it is first step since sse[0] = infnite
        if step >= 1:
            # Update the improvement in SSE
            error_improve = (sum_errors[step - 1] - sum_errors[step]) / sum_errors[step - 1]
        #    #Print result
        # print("Iteration {}: SSE = {:.4f}".format(step + 1, sum_errors[step]))

        # file.write("Iteration {}: SSE = {:.4f}".format(step+1,sum_errors[step])+'\n')

        # Assign the new centers for next iterations
        for i in range(clusters):
            centers[i] = data_cluster_errs[i, 3:data.shape[1] + 1] / data_cluster_errs[i, 2]
        # Exit the while loop if there is no improvement in SSE
        if error_improve == 0:
            # print("Final SSE = {:.4f}".format(sum_errors[step]))
            # print('Number of Iterations: {}'.format(step))
            return data[:,1]
        else:
            step += 1
    # print("Final SSE = {:.4f}".format(sum_errors[step - 1]))
    # print('Number of Iterations: {}'.format(step))
    # Return the lastest SSE
    #return sum_errors[step - 1]
    return data[:,1]
# Eduidance distance
def dist(a, b):
    distance = 0
    for i in range(a.shape[0]):
        distance += (a[i]-b[i])*(a[i]-b[i])

    return distance

def External_Meature(precited_labels, true_labels):
    num_points = len(precited_labels)
    tp =0
    ni = 0
    mj = 0
    COLS,true_counts = np.unique(true_labels, return_counts=True)
    COLS = len(COLS)
    ROWS, predicted_counts = np.unique(precited_labels, return_counts=True)
    ROWS = len(ROWS)
    results = np.zeros((ROWS, COLS))
    for i in range(num_points):
        results[int(precited_labels[i]), int(true_labels[i])] += 1
    for i in range(ROWS):
        mj += true_counts[i]*true_counts[i]
        ni += predicted_counts[i]*predicted_counts[i]
        for j in range(COLS):
            if results[i][j] != 0:
                tp += results[i][j] * results[i][j]
    tn = (num_points*num_points - ni -mj+tp)/2
    fn = (mj-tp)/2
    fp = (ni- tp)/2
    tp = (tp-num_points)/2

    Jaccard = tp/(tp+fn+fp)
    rand = (tp+tn)/ scipy.misc.comb(num_points,2)
    prec = tp/(tp+fp)
    recall = tp/(tp+fn)
    FM = np.sqrt(prec*recall)

    return Jaccard, rand, FM

#Get the best run of k mean with random centers
def Best_run_rdCenters(data, cluster, iteration, thresold,run, data_labels):
    # best initial SSEs, best final SSEs, and best # iterations
    best_Jaccard, best_rand, best_FM = K_Mean_rdCenters(data, cluster, iteration, thresold, data_labels)
    # More runs
    for i in range(run - 1):
        print("Run {}".format(i))
        run_Jaccard, run_rand, run_FM = K_Mean_rdCenters(data, cluster, iteration, thresold, data_labels)
        # Replace the best_run if other run has better result
        if  run_Jaccard > best_Jaccard:
            best_Jaccard = run_Jaccard
        if run_rand > best_rand:
            best_rand = run_rand
        if run_FM > best_FM:
            best_FM = run_FM
    return best_Jaccard, best_rand, best_FM


#
# def main():
#     #data = load_txt("D:\Python\DataClustering\Phase1\iris_bezdek.txt")
#     num_clusters, data, data_labels = Min_Max(r"D:\Python\DataClustering\Phase4\mfeat-fou.txt")
#     print(data.shape)
#     print(data_labels)
#     print(data)
#     # print(data_clusters)
#     # print(data_labels)
#
#
#
# main()
# import openpyxl module to open excel file to write data in
import openpyxl
if __name__ == "__main__":
#     #Getting the filename, number of clusters, iterations, threshold, and run from cmd
    print("Script name: %s" % str(sys.argv[0]))
    print("filename: %s" % str(sys.argv[1]))
    # print("clusters: %s" % int(sys.argv[2]))
    print("iterations: %s" % int(sys.argv[2]))
    print("thresold: %s" % float(sys.argv[3]))
    print("run: %s" % int(sys.argv[4]))
    print("Saved as: %s" % str(sys.argv[5]))
    saved_sheet = str(sys.argv[5])
    # file = open(saved_file, "w")
    # file.write("Script name: %s" % str(sys.argv[0]) +'\n')
    # file.write("Filename: %s" % str(sys.argv[1]) +'\n')
    # file.write("Clusters: %s" % int(sys.argv[2])+'\n')
    # file.write("Iterations: %s" % int(sys.argv[3])+'\n')
    # file.write("Threshold: %s" % float(sys.argv[4])+'\n')
    # file.write("Number of runs: %s" % int(sys.argv[5])+'\n')
    # file.write("Saved as: %s" % str(sys.argv[6])+'\n')
    wb = openpyxl.load_workbook(filename='D:\Python\DataClustering\Phase4\Phase4.xlsx')
    # Sheets can be added to workbook with the
    # workbook object's create_sheet() method.
    ws = wb[saved_sheet]
    ws['A1'] = 'Filename'
    ws['A2'] = 'Clusters'
    ws['A3'] = 'Max-Iterations'
    ws['A4'] = 'Thresold'
    ws['A5'] = 'Number of Run'

    ws['B1'] = str(sys.argv[1])
    ws['B3'] = int(sys.argv[2])
    ws['B4'] = float(sys.argv[3])
    ws['B5'] = int(sys.argv[4])

    ws['A6'] = 'Jaccard'
    ws['A7'] = 'Rand'
    ws['A8'] = 'FM'

    filename = str(sys.argv[1])
    run = int(sys.argv[4])
    num_interations = int(sys.argv[2])
    threshold = float(sys.argv[3])
    #Min-Max Normalization
    num_clusters, data, data_labels = Min_Max(filename)
    print("clusters: %s" % num_clusters)
    # Get the best run of k mean with random centers
    best_Jaccard, best_rand, best_FM = Best_run_rdCenters(data, num_clusters, num_interations, threshold,run, data_labels)
    ws['B2'] = num_clusters
    ws['B6'] = best_Jaccard
    ws['B7'] = best_rand
    ws['B8'] = best_FM


    wb.save('D:\Python\DataClustering\Phase4\Phase4.xlsx')