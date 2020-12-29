"""
This program is responding to Phase 2: Normalization and Initialization
Min_Max() is responding to Main-Max normalization
K_Mean_Maximin() is responding to K-Mean with Maximin method
K_mean() is responding to the general steps after center selecting, Initial SEE, Final SSE, Numver of iterations are updating here
DB() is responding to Davies–Bould
SW is responding to Silhouette Coefficient/Width
CH is responding to Calinski–Harabasz Index
"""
import pandas as pd
import random as rd
import sys
import math
import numpy as np



#Min-Max normalization
def Min_Max(filename):
    data = pd.read_csv(filename, sep=" ", skiprows=1, header=None)
    #v'=((v-min)/(max-min))*(new_max-new_min)+new_min
    #new_max = 1; new_min=0 here
    #so the equation will be v'=(v-min)/(max-min)
    #data.min() returns the minimum of each column
    #data.max() returns the maximum of each column
    data = (data-data.min())/(data.max()-data.min())
    data = data.fillna(0)
    data = data.values
    return data

#K-Mean with Maximin method
def K_Mean_Maximin(data, clusters, iterations, threshold):
    #create an empty array with NaN values
    centers  = np.empty((clusters,data.shape[1]))
    centers[:] = np.NaN
    #randomly select the first center
    center_first = rd.randrange(0,data.shape[0],1)
    centers[0] = data[center_first,:]
    #loop the all array that contains the centers
    for i in range(clusters):
        # we will fill up the center where is empty
        if(np.isnan(centers[i])).any() == True:
            # greatest distance
            dis_max = 0
            # to track the center
            track_center = 0
            #Look for the point that has the greatest edclidean distance to the nearest previously selected (i-1) centers
            for j in range(data.shape[0]):
                # Eduidance distance;
                # this loop will return the smallest distance between the point to each centers
                for k in range(i):
                    if k == 0:
                        dist_center_x_0 = dist(centers[k],data[j])
                    else:
                        dist_center_x = dist(centers[k],data[j])
                        if (dist_center_x < dist_center_x_0):
                            dist_center_x_0 = dist_center_x
                #assign the center with greatest distance among all
                if (dis_max < dist_center_x_0):
                    dis_max = dist_center_x_0
                    track_center = j
            centers[i] = data[track_center]
    # Add a column to keep track of the cluster of each data point; as fitst column of the dataset
    data = np.c_[[99] * data.shape[0], data]
    final_sse, mean,data_cluster_errs, centers, data_new = K_Mean(centers, data, clusters, iterations, threshold)
    return final_sse,mean,data_cluster_errs, centers, data_new

def K_Mean(centers,data, clusters, iterations, threshold):
    #Calculate mean of the whole dataset
    mean = np.mean(data, axis=0)
    # Create a list to store the SSE of each iterations
    sum_errors = [99] * iterations

    # #A value to store the improvement of SSE in each iterations
    error_improve = math.inf

    # Create the the 'distance' column to riginal data to indincate the distance of the instance to their cluster center, first column of the array
    data = np.c_[[math.inf] * data.shape[0], data]
    #
    # #A value to keep track of the iteration step (Which iteration it is)
    step = 0

    # #Let jump into the while loop
    # #The while loop will exit if the improvment of SSE small than threshold, or running out of the iterations, or there is no improvememnt in SSE
    while ((error_improve - threshold >= 0) and (iterations - step > 0)):
        # Reset the distance to infinite
        data[:, 0] = [math.inf] * data.shape[0]
        # Calculate all the distance between all the instances to all the cluster centers
        for i in range(data.shape[0]):
            for j in range(clusters):
                # Eduidance distance
                dis = dist(data[i, 2:data.shape[1]], centers[j, :])
                # Store the distance to the cluster_dis dataframe
                # Assign the instance to the closest cluster
                if (dis < data[i][0]):
                    # assign j to the cluster
                    data[i][1] = j
                    # assign closest distance
                    data[i][0] = dis

        data_cluster_errs = np.zeros((clusters, data.shape[1] + 1))
        for i in range(data.shape[0]):
            cluster = int(data[i, 1])
            data_cluster_errs[cluster, 0] += data[i, 0]
            data_cluster_errs[cluster, 1] = cluster
            data_cluster_errs[cluster, 2] += 1
            data_cluster_errs[cluster, 3:data.shape[1] + 1] += data[i, 2:data.shape[1]]

        # Handling empty cluster
        # Check if there is any 0 in the third column of data_cluster errs
        if (np.isin(0, data_cluster_errs[:, 2]).any()) == True:
            for i in range(clusters):
                # Pass if the cluster is not empty(the count is not 0)
                if (np.isin(0, data_cluster_errs[i, 2]).all()) == False:
                    pass
                else:
                    # Sort the data based on the distance with ascend order, then get the last index values, which contributes the most to the overall SSE
                    data = data[np.argsort(data[:, 0])]
                    # Change the cluster name to be the empty cluster
                    data[-1, 1] = i
                    # Change the distance to center to 0
                    data[-1, 0] = 0
            data_cluster_errs = np.zeros((clusters, data.shape[1] + 1))
            for i in range(data.shape[0]):
                cluster = int(data[i, 1])
                data_cluster_errs[cluster, 0] += data[i, 0]
                data_cluster_errs[cluster, 1] = cluster
                data_cluster_errs[cluster, 2] += 1
                data_cluster_errs[cluster, 3:data.shape[1] + 1] += data[i, 2:data.shape[1]]
        # Get the sum of Sum-of-Squared Error
        sum_errors[step] = data_cluster_errs[:, 0].sum()

        # if step == 0:
            # Get the initial SSE
            # print("Initial SSE = {:.4f}".format(sum_errors[step])+"; " +"delta SSE = 0")
            # Skip this part if it is first step since sse[0] = infnite
        if step >= 1:
            # Update the improvement in SSE
            error_improve = (sum_errors[step - 1] - sum_errors[step]) / sum_errors[step - 1]
        #    #Print result
        #     print("Iteration {}: SSE = {:.4f}".format(step + 1, sum_errors[step]) + "; " +"delta SSE = {:}".format(error_improve))

        # file.write("Iteration {}: SSE = {:.4f}".format(step+1,sum_errors[step])+'\n')

        # Exit the while loop if there is no improvement in SSE
        if error_improve == 0:
            # print("Final SSE = {:.4f}".format(sum_errors[step]))
            # print('Number of Iterations: {}'.format(step))
            # ch_k = CH(mean, data_cluster_errs[:, 1:3], centers, sum_errors[step - 1])
            # sw_k = SW(data, centers,data_cluster_errs[:,2])
            # db_k = DB(data, data_cluster_errs)
            return sum_errors[step], mean,data_cluster_errs, centers, data
        else:
            # Assign the new centers for next iterations
            for i in range(clusters):
                centers[i] = data_cluster_errs[i, 3:data.shape[1] + 1] / data_cluster_errs[i, 2]
            step += 1
    # print("Final SSE = {:.4f}".format(sum_errors[step - 1]))
    # print('Number of Iterations: {}'.format(step))
    # Return the lastest SSE
    #return sum_errors[step - 1]
    # ch_k = CH(mean,data_cluster_errs[:,1:3], centers, sum_errors[step - 1])
    # sw_k = SW(data, centers,data_cluster_errs[:,2])
    # db_k = DB(data, data_cluster_errs)
    return sum_errors[step - 1], mean,data_cluster_errs, centers, data
#Davies–Bould
def DB(data, data_cluster_errs):
    means = np.array([data_cluster_errs[i, 3:data.shape[1] + 1] / data_cluster_errs[i, 2] for i in range(data_cluster_errs.shape[0])])
    stds = np.zeros((data_cluster_errs.shape[0],))
    for i in range(data.shape[0]):
        for j in range(data_cluster_errs.shape[0]):
            if data[i,1] == j:
                stds[j] += dist(data[i, 2:data.shape[0]], means[j,:])*dist(data[i, 2:data.shape[0]], means[j,:])
    for counts in range(stds.shape[0]):
        stds[counts] =  np.sqrt(stds[counts] / data_cluster_errs[counts, 2])

    dbs = 0
    for i in range(stds.shape[0]):
        db = 0
        for j in range(0, stds.shape[0]):
            if(i != j):
                db_i = (stds[i]+stds[j])/(dist(means[i],means[j]))
                if (db_i>db):
                    db = db_i
        dbs += db
    dbs = dbs/stds.shape[0]

    return dbs


# #Silhouette Coefficient/Width
def SW(data,centers,cluster_count):
    # Add a column to keep track of the mean distance from xi to points in its own cluster y.
    data = np.c_[[0] * data.shape[0], data]
    # Add a column to keep track of the mean of the distances from xi to points in the closest cluster
    data = np.c_[[0] * data.shape[0], data]

    # print(data[:,0])
    # data = data[np.argsort(data[:,3])]
    # print(data[:,3])
    # print(data[0])
    # print(data[0, 4:data.shape[1]])
    #cluster_counts = dict(zip(unique, counts))
    # print(cluster_counts)
    for i in range(data.shape[0]):
        dis_ic = math.inf
        closest_cluster = 0
        for c in range(centers.shape[0]):
            dis = dist(data[i, 4:data.shape[1]], centers[c])
            if c!= data[i,3]:
                if dis < dis_ic:
                    dis_ic = dis
                    closest_cluster = c
        for j in range(data.shape[0]):
            if i != j:
                if data[i,3] == data[j,3]:
                    data[i,1] += dist(data[i, 4:data.shape[1]],data[j, 4:data.shape[1]])
                if data[j,3] == closest_cluster:
                    data[i,0] += dist(data[i, 4:data.shape[1]],data[j, 4:data.shape[1]])
        data[i,0]= data[i,0]/cluster_count[closest_cluster]
        data[i,1] = data[i,1]/cluster_count[int(data[i,3])]


    # print(data[:,0])
    # print(data[:,1])
    si = 0
    for i in range(data.shape[0]):
        si += (data[i,0]- data[i,1])/max(data[i,0], data[i,1])
    si = si/data.shape[0]

    return si

#Calinski–Harabasz Index
def CH(data_mean, cluster_count, centers, final_sse):
    data_mean = data_mean[1: len(data_mean)]
    num_points = np.sum(cluster_count[:,1])
    sb_sum = 0
    for i in range(cluster_count.shape[0]):
        sb_sum += cluster_count[i, 1] * np.dot((centers[i] - data_mean), (centers[i] - data_mean).T)
    ch_k = ((num_points-cluster_count.shape[0])*sb_sum)/((cluster_count.shape[0]-1)*final_sse)

    return ch_k


# Eduidance distance
def dist(a, b):
    distance = 0
    for i in range(a.shape[0]):
        distance += (a[i]-b[i])*(a[i]-b[i])

    return distance


#Get the best run of k mean with maximin method
def Best_run_maxmin(data, cluster, iteration, thresold,run):
    # best initial SSEs, best final SSEs, and best # iterations
    final_sse,  mean,data_cluster_errs, centers, data_new = K_Mean_Maximin(data, cluster, iteration, thresold)
    # More runs
    for i in range(run - 1):
        run_final_sse, run_mean,run_data_cluster_errs, run_centers, run_data_new = K_Mean_Maximin(data, cluster, iteration, thresold)
        if(run_final_sse<final_sse):
            final_sse = run_final_sse
            mean = run_mean
            data_cluster_errs = run_data_cluster_errs
            centers = run_centers
            data_new = run_data_new
    ch_k = CH(mean,data_cluster_errs[:,1:3], centers,final_sse)
    sw_k = SW(data_new, centers,data_cluster_errs[:,2])
    db_k = DB(data_new, data_cluster_errs)

    return ch_k, sw_k, db_k
# import openpyxl
# def main():
#     data = Min_Max("D:\Python\DataClustering\Phase1\iris_bezdek.txt")
#     k_min = 2
#     k_max = np.round(np.sqrt((data.shape[0])/2),0)
#     # wb = openpyxl.load_workbook(filename='D:\Python\DataClustering\Phase3\Phase3.xlsx')
#     # Sheets can be added to workbook with the
#     # workbook object's create_sheet() method.
#     # ws = wb['iris_bezdek_Result']
#     # ws['A1'] = 'Filename'
#     # ws['A2'] = 'Max-Iterations'
#     # ws['A3'] = 'Thresold'
#     # ws['A4'] = 'Number of Run'
#     # ws['A5'] = "Clusters"
#     # ws['B6'] = "CH"
#     # ws['B7'] = "SW"
#     # ws['B8'] = "DB"
#     num_interations = 100
#     threshold = 0.001
#     run = 100
#     CHs = np.zeros((int(k_max-k_min+1),))
#     SWs = np.zeros((int(k_max - k_min + 1),))
#     DBs = np.zeros((int(k_max - k_min + 1),))
#     #
#     for i in range((int(k_max-k_min+1))):
#         CHs[i], SWs[i], DBs[i] = Best_run_maxmin(data, i+2, num_interations, threshold,run)
#         # ws.cell(row=5, column=i + 3).value = "K = {:}".format(i + 2)
#         # ws.cell(row=6, column=i + 3).value = CHs[i]
#         # ws.cell(row=7, column=i + 3).value = SWs[i]
#         # ws.cell(row=8, column=i + 3).value = DBs[i]
#         print("CH({:}) = {:}".format(i+2, CHs[i]))
#         print("SW({:}) = {:}".format(i + 2, SWs[i]))
#         print("DB({:}) = {:}".format(i + 2, DBs[i]))
#
#     # wb.save('D:\Python\DataClustering\Phase3\Phase3.xlsx')
#
#
#
# main()
#import openpyxl module to open excel file to write data in
import openpyxl
if __name__ == "__main__":
#     #Getting the filename, number of clusters, iterations, threshold, and run from cmd
    print("Script name: %s" % str(sys.argv[0]))
    print("filename: %s" % str(sys.argv[1]))
    #print("clusters: %s" % int(sys.argv[2]))
    print("iterations: %s" % int(sys.argv[2]))
    print("thresold: %s" % float(sys.argv[3]))
    print("run: %s" % int(sys.argv[4]))
    print("Saved as: %s" % str(sys.argv[5]))
    saved_sheet = str(sys.argv[5])
    # data = Min_Max("D:\Python\DataClustering\Phase1\iris_bezdek.txt")

    wb = openpyxl.load_workbook(filename='D:\Python\DataClustering\Phase3\Phase3.xlsx')
    # Sheets can be added to workbook with the
    # workbook object's create_sheet() method.
    ws = wb[saved_sheet]
    ws['A1'] = 'Filename'
    ws['A2'] = 'Max-Iterations'
    ws['A3'] = 'Thresold'
    ws['A4'] = 'Number of Run'
    ws['A5'] = "Clusters"
    ws['B6'] = "CH"
    ws['B7'] = "SW"
    ws['B8'] = "DB"
    # ws['B9'] = "Final SSE"
    ws['B1'] = str(sys.argv[1])
    ws['B2'] = int(sys.argv[2])
    ws['B3'] = float(sys.argv[3])
    ws['B4'] = int(sys.argv[4])

    filename = str(sys.argv[1])
    num_interations = int(sys.argv[2])
    threshold = float(sys.argv[3])
    run = int(sys.argv[4])
    #Min-Max Normalization
    data = Min_Max(filename)
    k_min = 2
    k_max = np.round(np.sqrt((data.shape[0]) / 2), 0)
    # final_sses= np.zeros((int(k_max - k_min + 1),))
    CHs = np.zeros((int(k_max - k_min + 1),))
    SWs = np.zeros((int(k_max - k_min + 1),))
    DBs = np.zeros((int(k_max - k_min + 1),))
    #
    for i in range((int(k_max - k_min + 1))):
        CHs[i], SWs[i], DBs[i] = Best_run_maxmin(data, i + 2, num_interations, threshold, run)
        ws.cell(row=5, column=i + 3).value = "K = {:}".format(i + 2)
        ws.cell(row=6, column=i + 3).value = CHs[i]
        ws.cell(row=7, column=i + 3).value = SWs[i]
        ws.cell(row=8, column=i + 3).value = DBs[i]
        # ws.cell(row=9, column=i + 3).value = final_sses[i]
        print("CH({:}) = {:}".format(i + 2, CHs[i]))
        print("SW({:}) = {:}".format(i + 2, SWs[i]))
        print("DB({:}) = {:}".format(i + 2, DBs[i]))
        # print("Final SSE({:}) = {:}".format(i + 2, final_sses[i]))

    wb.save('D:\Python\DataClustering\Phase3\Phase3.xlsx')